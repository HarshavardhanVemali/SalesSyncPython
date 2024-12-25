from django.shortcuts import render,redirect
from django.http import JsonResponse,HttpResponse
from django.contrib.auth import authenticate, login,logout
from .models import FailedLoginAttempts
import uuid
import json
from django.views.decorators.csrf import csrf_exempt,csrf_protect
from django.contrib.auth.models import User
from .models import Customer, Order, Payment
from datetime import datetime,date
from django.core.exceptions import ObjectDoesNotExist
from decimal import Decimal, InvalidOperation
from django.db import transaction
from reportlab.pdfgen import canvas
from reportlab.lib.pagesizes import landscape, letter
from reportlab.lib.units import inch, mm,cm
from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph
from reportlab.lib import colors
from reportlab.lib.styles import getSampleStyleSheet
from django.db.models import Sum
import heapq
from reportlab.lib.pagesizes import A3
import xlwt
from django.views.decorators.csrf import ensure_csrf_cookie
from django.middleware.csrf import get_token
from django.contrib.auth.models import Permission, ContentType
from django.contrib.auth.decorators import user_passes_test
from openpyxl import load_workbook
from openpyxl.utils import datetime as opxldatetime
from django.utils import timezone
from rest_framework_simplejwt.tokens import RefreshToken

def get_token(request):
    
    if 'access_token' in request.COOKIES:
      return request.COOKIES.get('access_token')
    return None


def admin_required(view_func):
    return user_passes_test(
        lambda u: u.is_superuser or u.is_staff,
        login_url='/adminlogin/'
    )(view_func)

def user_required(view_func):
    return user_passes_test(
        lambda u: not u.is_superuser and not u.is_staff,
        login_url='/userlogin/'
    )(view_func)

def index(request):
    return render(request,'index.html')

MAX_FAILED_ATTEMPTS = 3

def set_device_cookie(response, device_id):
    response.set_cookie(
        'device_id', 
        device_id, 
        max_age=365 * 24 * 60 * 60, 
        httponly=True, 
        samesite='Lax'  
    )

def get_device_id(request):
    if 'device_id' in request.COOKIES:
        return request.COOKIES['device_id']
    return str(uuid.uuid4())

def is_device_blocked(device_id):
    try:
        failed_attempt = FailedLoginAttempts.objects.get(device_id=device_id)
        if failed_attempt.is_active:
            return True
    except FailedLoginAttempts.DoesNotExist:
        return False
    return False

def get_tokens_for_user(user):
    refresh = RefreshToken.for_user(user)

    return {
        'refresh': str(refresh),
        'access': str(refresh.access_token),
    }


@csrf_exempt
def adminlogin(request):
    response = HttpResponse() 
    if 'device_id' not in request.COOKIES:  
        device_id = get_device_id(request)
        set_device_cookie(response, device_id)

    if request.method == 'POST':
        device_id = request.COOKIES.get('device_id') 
        if is_device_blocked(device_id):
            return JsonResponse({'blocked': True, 'error_message': 'Device is permanently blocked.'}, status=400)
        data=json.loads(request.body)
        username = data.get('username')
        password = data.get('password')
        user = authenticate(request, username=username, password=password)

        if user is not None:
            if user.is_superuser or user.is_staff and user.is_active:
                FailedLoginAttempts.objects.filter(device_id=device_id).update(attempts=0, is_active=False)
                login(request, user)
                print('logged')
                tokens = get_tokens_for_user(user)

                response = JsonResponse({'success': True, 'access_token': tokens['access'], 'refresh_token': tokens['refresh']}, status=200)
                response.set_cookie('access_token',tokens['access'],httponly=True)
                response.set_cookie('refresh_token',tokens['refresh'],httponly=True)

                return response
            return JsonResponse({'error_message': 'Permissions Denied'}, status=400)
        failed_attempt, created = FailedLoginAttempts.objects.get_or_create(device_id=device_id)
        if not created:
            failed_attempt.attempts += 1
            failed_attempt.save()
            if failed_attempt.attempts >= MAX_FAILED_ATTEMPTS:
                failed_attempt.is_active = True
                failed_attempt.save()
                return JsonResponse({'error_message': 'Device is permanently blocked'}, status=400)
        else:
            failed_attempt.attempts = 1
            failed_attempt.save()

        return JsonResponse({'error_message': 'Invalid username or password'}, status=400)
    else:  
        if 'device_id' not in request.COOKIES:
            device_id = get_device_id(request)
            response = JsonResponse({'error_message': 'Invalid request.'})
            set_device_cookie(response, device_id)
            return response

        return JsonResponse({'error_message': 'Invalid request.'})


@csrf_exempt
def logout_admin(request):
    logout(request)
    response = JsonResponse({'message': 'Successfully logout'}, status=200)
    response.delete_cookie('access_token')
    response.delete_cookie('refresh_token')
    return response


@csrf_exempt
def adduser(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            username = data.get('username')
            useremail = data.get('useremail')
            usertype = data.get('usertype')
            if User.objects.filter(email=useremail).exists():
                return JsonResponse({'error_userexits_message': 'User already exists.'})
            if usertype == 'admin':
                create_user = User.objects.create_superuser(first_name=username, username=useremail,email=useremail, password='defaultpassword')
            elif usertype == 'view':
                create_user = User.objects.create_user(first_name=username, username=useremail,email=useremail, password='defaultpassword')
            else:
                return JsonResponse({'error_message': 'Invalid user type.'}, status=400)

            return JsonResponse({'success': True})
        except Exception as e:
            return JsonResponse({'error_message': str(e)}, status=500)

    return JsonResponse({'error_message': 'Invalid request method.'}, status=405)

@csrf_exempt
def getuserdetails(request):
    if request.method == 'GET':
        try:
            users = User.objects.values(
                'id', 'first_name', 'email', 'is_staff', 'is_superuser', 'is_active' 
            )
            user_list = list(users) 
            for user in user_list: 
                if user['is_superuser']: 
                    user['user_type'] = 'Admin'
                else:
                    user['user_type'] = 'View'
                del user['is_superuser'] 
                del user['is_staff']
            return JsonResponse({'users': user_list}, status=200, safe=False)
        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)
    return JsonResponse({'error_message':'Invalid request.'})

@csrf_exempt
def resetpassword(request):
    if request.method=='POST':
        data=json.loads(request.body)
        email=data.get('userId')
        if not email:
            return JsonResponse({'error_message':'Required missing fields.'},status=400)
        try:
            user=User.objects.filter(email=email)
            if user:
                user.password='defaultpassword'
                return JsonResponse({'success':True})
            else:
                return({'error_message':'Invalid.'})
        except Exception as e:
            return JsonResponse({'error_message':str(e)})
    return JsonResponse({'error_message':'Invalid request'},status=500)

@csrf_exempt
def updateusertype(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_id = data.get('userId') 
            new_type = data.get('newType')
            try:
                user = User.objects.get(email=user_id) 
            except User.DoesNotExist:
                return JsonResponse({'error_message': 'User not found.'}, status=404)
            if new_type == 'admin':
                user.is_superuser = True
                user.is_staff = True 
            elif new_type == 'view':
                user.is_superuser = False
                user.is_staff = False 
            else:
                return JsonResponse({'error_message': 'Invalid user type.'}, status=400)
            user.save()
            return JsonResponse({'message': 'User type updated successfully.'}, status=200)
        except json.JSONDecodeError:
            return JsonResponse({'error_message': 'Invalid JSON data.'}, status=400)
        except Exception as e:
            return JsonResponse({'error_message': str(e)}, status=500)
    else:
        return JsonResponse({'error_message': 'Method Not Allowed.'}, status=405)
    
@csrf_exempt
def deleteusers(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            user_ids = data.get('user_ids')
            if not user_ids:
                return JsonResponse({'error': 'No user IDs provided.'}, status=400)
            if not isinstance(user_ids, list):
                return JsonResponse({'error': 'Invalid user IDs format.'}, status=400)
            deleted_count, _ = User.objects.filter(email__in=user_ids).delete()
            if deleted_count > 0:
                return JsonResponse({'message': 'Users deleted successfully.'}, status=200)
            else:
                return JsonResponse({'message': 'User not found.'}, status=404) 
        except json.JSONDecodeError as e:
            return JsonResponse({'error': 'Invalid JSON data: ' + str(e)}, status=400)  

        except Exception as e:
            return JsonResponse({'error': 'An unexpected error occurred: ' + str(e)}, status=500)  

    else:
        return JsonResponse({'error': 'Method Not Allowed.'}, status=405)  


@csrf_exempt
def addreport(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            customername = data.get('customername')
            customermobile = data.get('customermobile')
            customerarea = data.get('customerarea')
            product = data.get('product')
            dateofinvoice_str = data.get('dateofinvoice') 
            amount_str = data.get('amount') 
            amountpaid_str = data.get('amountpaid') 
            remarks = data.get('remarks')
            paymentmethod = data.get('paymentmethod')

            required_fields = ['customername', 'customermobile', 'customerarea', 'product', 'dateofinvoice', 'amount', 'amountpaid', 'remarks', 'paymentmethod']
            missing_fields = [field for field in required_fields if not data.get(field)]
            if missing_fields:
                return JsonResponse({'errors': {'Missing required fields: '}}, status=400)
            amount = float(amount_str)
            amountpaid = float(amountpaid_str)
            dateofinvoice = datetime.strptime(dateofinvoice_str, '%Y-%m-%d').date()

            customer, created = Customer.objects.get_or_create(phone=customermobile, defaults={'name': customername, 'area': customerarea})

            order = Order.objects.create(
                customer=customer,
                product=product,
                billing_date=dateofinvoice,
                amount=amount,
            )
            Payment.objects.create(
                order=order,
                payment_date=dateofinvoice,
                paid_amount=amountpaid,
                payment_method=paymentmethod,
                remarks=remarks
            )
            return JsonResponse({'success': True, 'message': 'Report added successfully'}, status=201)

        except (KeyError, ValueError, ObjectDoesNotExist, Exception) as e:
            return JsonResponse({'errors': {'general': str(e)}}, status=400)  
    else:
        return JsonResponse({'error_message': 'Invalid request'}, status=405) 

@csrf_exempt
def getreports(request):
    if request.method == 'GET':
        try:
            orders = Order.objects.all().prefetch_related('payment_set')
            order_data = []
            status_map = {
                'partially_paid': 'Partially Paid',
                'paid': 'Paid',
                'unpaid': 'Unpaid',
                'cash': 'Cash',
                'card':'Card',
                'cheque':'Cheque',
                'upi':'UPI',
                'bank_transfer':'Bank Transfer'
            }
            for order in orders:
                payments = order.payment_set.all()
                payment_details = []
                for payment in payments:
                    payment_details.append({
                        'payment_date': payment.payment_date.strftime('%d/%m/%Y'),
                        'paid_amount': payment.paid_amount,
                        'payment_method':status_map.get(payment.payment_method, 'Unknown'),
                        'remarks': payment.remarks,
                    })
                order_data.append({
                    'bill_no': order.bill_no,
                    'customer_name': order.customer.name,
                    'mobile_number': order.customer.phone if order.customer.phone else '-',
                    'area': order.customer.area,
                    'product': order.product,
                    'amount': order.amount,
                    'status': status_map.get(order.status, 'Unknown Status'),
                    'billing_date': order.billing_date.strftime('%d/%m/%Y'),
                    'payments': payment_details, 
                    'total_paid': order.total_paid, 
                    'balance': order.balance,
                })
            return JsonResponse({'orders': order_data}, safe=False) 

        except Exception as e:
            return JsonResponse({'error': str(e)}, status=500)

@csrf_exempt
def addpayment(request):
    if request.method == 'POST':
        try:
            with transaction.atomic():
                data = json.loads(request.body)
                bill_no = data.get('bill_no')
                payment_date = data.get('payment_date')
                paid_amount_str = data.get('paid_amount') 
                payment_method = data.get('payment_method')
                remarks = data.get('remarks')
                required_fields = [bill_no, payment_method, paid_amount_str, payment_date, remarks]
                if not all(required_fields):
                    return JsonResponse({'error_message': 'Required fields are missing.'}, status=400)

                try:
                    paid_amount = Decimal(paid_amount_str) 
                except InvalidOperation as e:
                    return JsonResponse({'error_message': 'Invalid paid_amount format.'}, status=400)

                try:
                    order = Order.objects.get(bill_no=bill_no) 
                except Order.DoesNotExist:
                    return JsonResponse({'error_message': 'Order does not exist.'}, status=400)

                payment = Payment.objects.create(
                    order=order,
                    payment_date=payment_date,
                    paid_amount=paid_amount,
                    payment_method=payment_method,
                    remarks=remarks,
                )
                return JsonResponse({'success': True}, status=200)
        except Exception as e:
            return JsonResponse({'error_message': str(e)}, status=500)
    return JsonResponse({'error_message': 'Invalid request method.'}, status=405) 

@csrf_exempt
def generateledger(request):
    if request.method == 'POST':
        try:
            data = json.loads(request.body)
            customer_mobile = data.get('customermobile')
            start_date_str = data.get('from')
            end_date_str = data.get('to')
            company_name = customer_mobile

            if not all([customer_mobile, start_date_str, end_date_str, company_name]):
                return JsonResponse({'error_message': 'Required fields are missing.'}, status=400)

            try:
                customer = Customer.objects.get(name=customer_mobile)
            except Customer.DoesNotExist:
                return JsonResponse({'error_message': 'Customer not found.'}, status=404)
            company_address=customer.area
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError:
                return JsonResponse({'error_message': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

            opening_balance = (Order.objects.filter(customer=customer, billing_date__lt=start_date)
                               .aggregate(total_amount=Sum('amount'))['total_amount'] or Decimal(0.00))
            opening_balance -= (Payment.objects.filter(order__customer=customer, payment_date__lt=start_date)
                                .aggregate(total_paid=Sum('paid_amount'))['total_paid'] or Decimal(0.00))

            ledger_entries = []
            
            orders = Order.objects.filter(customer=customer, billing_date__range=(start_date, end_date)).order_by('billing_date')
            payments = Payment.objects.filter(order__customer=customer, payment_date__range=(start_date, end_date)).order_by('payment_date')
            
            all_entries = []
            for order in orders:
                all_entries.append({
                    "Doc.No": order.bill_no,
                    "Doc.Date": order.billing_date,
                    "Doc.Type": "RV",
                    "Item.Texts": f"Sales Invoice - {order.product}/Posting Dt.{order.billing_date}",
                    "Dr.Amount": order.amount,
                    "Cr.Amount": Decimal(0.00),
                    "Balance": None,
                    "date": order.billing_date  
                })
            for payment in payments:
                all_entries.append({
                    "Doc.No": payment.order.bill_no,
                    "Doc.Date": payment.payment_date,
                    "Doc.Type": "DZ",
                    "Item.Texts": f"Payment - {payment.remarks or 'No Remarks'}-{payment.order.bill_no}/Posting Dt.{payment.payment_date}",
                    "Dr.Amount": Decimal(0.00),
                    "Cr.Amount": payment.paid_amount,
                    "Balance": None,
                    "date": payment.payment_date  
                })
            
            all_entries = sorted(all_entries, key=lambda x: x['date'])

            running_balance = opening_balance
            ledger_entries.append({
                "Doc.No": "",
                "Doc.Date": "",
                "Doc.Type": "",
                "Item.Texts": "Opening Balance",
                "Dr.Amount":  0,
                "Cr.Amount": 0,
                "Balance": opening_balance
            })

            for entry in all_entries:
                entry['Balance'] = running_balance
                if entry["Dr.Amount"]:
                    running_balance += entry["Dr.Amount"]
                if entry["Cr.Amount"]:
                    running_balance -= entry["Cr.Amount"]
                entry["Balance"] = running_balance
                ledger_entries.append(entry)
            scaling_factor = 1.4
            today=date.today()
            response = HttpResponse(content_type='application/pdf')
            response['Content-Disposition'] = f'attachment; filename="{company_name}_{today}_ledger.pdf"'
            c = canvas.Canvas(response, pagesize=landscape(A3))
            width, height = landscape(A3)

            margin_left = 10 * mm * scaling_factor
            margin_right = 15 * mm * scaling_factor
            margin_top = 25 * mm * scaling_factor
            top_section_height = 50 * mm * scaling_factor
            table_header_height = 15 * mm * scaling_factor
            line_height = 12 * scaling_factor

            c.setFont("Helvetica-Bold", 14 * scaling_factor)
            c.drawString(margin_left, height - margin_top + 15, "Aswini Associates")
            c.setFont("Helvetica", 8 * scaling_factor)
            c.drawString(margin_left, height - margin_top, "Regional Office - South \"Cavainville\", No-12, Cenotaph Road, Chennai-600018")

            c.line(margin_left, height - margin_top - 5, width - margin_left, height - margin_top - 5)

            company_info_y = height - margin_top - top_section_height + 120*scaling_factor
            c.setFont("Helvetica-Bold", 15 * scaling_factor)
            c.drawString(margin_left, company_info_y, "Company")
            c.setFont("Helvetica-Bold", 13 * scaling_factor)
            c.drawString(margin_left, company_info_y-25, company_name)
            c.setFont("Courier", 12 * scaling_factor)
            c.drawString(margin_left, company_info_y - 50, company_address)
            c.setFont("Courier", 14 * scaling_factor)
            c.drawString(margin_left, company_info_y-100, f"Account Statement from {start_date} to {end_date}")

            account_statement_width = 300 * scaling_factor
            account_statement_height = 80 * scaling_factor
            account_statement_x = width - margin_left - account_statement_width 
            account_statement_y = height - margin_top - account_statement_height-15

            c.setStrokeColor(colors.black) 
            c.setLineWidth(1)             
            c.rect(
                account_statement_x,
                account_statement_y,
                account_statement_width,
                account_statement_height, 
                stroke=1  
            )
            c.setFillColor(colors.lightgrey)
            c.rect(
                account_statement_x,
                account_statement_y + account_statement_height - 25,
                account_statement_width,
                25 * scaling_factor,
                fill=1,
            )
            text_margin = 10 * scaling_factor
            c.setFillColor(colors.black)
            c.setFont("Helvetica-Bold", 14 * scaling_factor)
            c.drawString(account_statement_x + text_margin, account_statement_y + account_statement_height - 15, "Account Statement")
            c.setFont("Helvetica-Bold", 14 * scaling_factor)

            c.drawString(account_statement_x + text_margin, account_statement_y + account_statement_height - 45, "Date:")
            c.setFont("Courier", 12 * scaling_factor)
            c.drawString(account_statement_x + text_margin + 70, account_statement_y + account_statement_height - 45, f"{end_date.strftime('%d.%m.%Y')}")
            c.setFont("Helvetica-Bold", 14 * scaling_factor)
            c.drawString(account_statement_x + text_margin, account_statement_y + account_statement_height - 75, "Your account with us")
            c.setFont("Courier", 12 * scaling_factor)
            c.drawString(account_statement_x + text_margin, account_statement_y + account_statement_height - 100, "2014494")

            current_y = height - margin_top - top_section_height + 20
            c.line(margin_left, current_y + 20, width - margin_left, current_y + 20)
            c.setFont("Courier-Bold", 12 * scaling_factor)

            def draw_page_border(c, margin_left, margin_right, margin_top, height):
                c.line(margin_left, margin_top - 15 * scaling_factor, width - margin_left, margin_top - 15 * scaling_factor)
            
            headers = ["Doc.No", "Doc.Date", "Doc.Type", "Item.Texts", "Dr.Amount", "Cr.Amount", "Balance"]
            header_x_positions = [margin_left, margin_left + 80 * scaling_factor, margin_left + 160 * scaling_factor, margin_left + 230 * scaling_factor, margin_left + 580 * scaling_factor, margin_left + 670 * scaling_factor, margin_left + 740 * scaling_factor]
            c.setFont("Courier-Bold", 12 * scaling_factor)
            for i, header in enumerate(headers):
                if header in ["Dr.Amount", "Cr.Amount", "Balance"]:
                    c.drawRightString(header_x_positions[i] + 40 * scaling_factor, current_y, header)
                else:
                    c.drawString(header_x_positions[i], current_y, header)
            current_y -= line_height

            c.setStrokeColor(colors.black)
            c.setLineWidth(1)
            c.line(margin_left, current_y , width - margin_left, current_y )
            current_y -= line_height

            c.setFont("Helvetica", 9 * scaling_factor)
            for entry in ledger_entries:
                if current_y < margin_top + line_height * 2: 
                    c.showPage()
                    current_y = height - margin_top - top_section_height
                    c.setFont("Courier-Bold", 12 * scaling_factor)
                    for i, header in enumerate(headers):
                        if header in ["Dr.Amount", "Cr.Amount", "Balance"]:
                            c.drawRightString(header_x_positions[i] + 40 * scaling_factor, current_y, header)
                        else:
                            c.drawString(header_x_positions[i], current_y, header)
                    current_y -= line_height
                    c.line(margin_left, current_y , width - margin_left, current_y )
                    current_y -= line_height
                    c.setFont("Helvetica", 9 * scaling_factor)

                entry_values = [
                    entry["Doc.No"],
                    entry["Doc.Date"].strftime('%d.%m.%Y') if entry["Doc.Date"] else '',
                    entry["Doc.Type"],
                    entry["Item.Texts"],
                    f"{entry['Dr.Amount']:.2f}" if entry['Dr.Amount'] > 0 else '0.00',
                    f"{entry['Cr.Amount']:.2f}" if entry['Cr.Amount'] > 0 else '0.00',
                    f"{entry['Balance']:.2f}"
                ]
                for i, value in enumerate(entry_values):
                    if headers[i] in ["Dr.Amount", "Cr.Amount", "Balance"]:
                        c.drawRightString(header_x_positions[i] + 40 * scaling_factor, current_y, str(value))
                    else:
                        c.drawString(header_x_positions[i], current_y, str(value))
                current_y -= line_height

            c.line(margin_left, margin_top + 12 * scaling_factor, width - margin_left, margin_top + 12 * scaling_factor)
            c.setFont("Courier", 12 * scaling_factor)
            balance_text = f"Final Balance as of {end_date.strftime('%d.%m.%Y')}:"
            balance_amount = f"{ledger_entries[-1]['Balance']:.2f}"
            text_y = margin_top - (line_height // 2) * scaling_factor
            c.drawString(margin_left, text_y, balance_text)
            c.drawRightString(width - margin_right - 10 * scaling_factor, text_y, balance_amount)

            draw_page_border(c, margin_left, margin_right, margin_top, height)

            c.save()
            return response
            

        except Exception as e:
            return JsonResponse({'error_message': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method.'}, status=405)

@csrf_exempt
def generate_excel_ledger(request):
    if request.method == 'POST':  
        try:
            data = json.loads(request.body)
            customer_mobile = data.get('customermobile')
            start_date_str = data.get('from')
            end_date_str = data.get('to')
            company_name = customer_mobile
            if not all([customer_mobile, start_date_str, end_date_str, company_name, company_address]):
                return JsonResponse({'error_message': 'Required fields are missing.'}, status=400)

            try:
                customer = Customer.objects.get(name=customer_mobile)
            except Customer.DoesNotExist:
                return JsonResponse({'error_message': 'Customer not found.'}, status=404)
            company_address=customer.area
            try:
                start_date = datetime.strptime(start_date_str, "%Y-%m-%d").date()
                end_date = datetime.strptime(end_date_str, "%Y-%m-%d").date()
            except ValueError:
                return JsonResponse({'error_message': 'Invalid date format. Use YYYY-MM-DD.'}, status=400)

            opening_balance = (Order.objects.filter(customer=customer, billing_date__lt=start_date)
                               .aggregate(total_amount=Sum('amount'))['total_amount'] or Decimal(0.00))
            opening_balance -= (Payment.objects.filter(order__customer=customer, payment_date__lt=start_date)
                                .aggregate(total_paid=Sum('paid_amount'))['total_paid'] or Decimal(0.00))

            ledger_entries = []
            
            orders = Order.objects.filter(customer=customer, billing_date__range=(start_date, end_date)).order_by('billing_date')
            payments = Payment.objects.filter(order__customer=customer, payment_date__range=(start_date, end_date)).order_by('payment_date')
            
            all_entries = []
            for order in orders:
                all_entries.append({
                    "Doc.No": order.bill_no,
                    "Doc.Date": order.billing_date,
                    "Doc.Type": "RV",
                    "Item.Texts": f"Sales Invoice - {order.product}/Posting Dt.{order.billing_date}",
                    "Dr.Amount": order.amount,
                    "Cr.Amount": Decimal(0.00),
                    "Balance": None,
                    "date": order.billing_date  
                })
            for payment in payments:
                all_entries.append({
                    "Doc.No": payment.order.bill_no,
                    "Doc.Date": payment.payment_date,
                    "Doc.Type": "DZ",
                    "Item.Texts": f"Payment - {payment.remarks or 'No Remarks'}-{payment.order.bill_no}/Posting Dt.{payment.payment_date}",
                    "Dr.Amount": Decimal(0.00),
                    "Cr.Amount": payment.paid_amount,
                    "Balance": None,
                    "date": payment.payment_date  
                })
            
            all_entries = sorted(all_entries, key=lambda x: x['date'])

            running_balance = opening_balance
            ledger_entries.append({
                "Doc.No": "",
                "Doc.Date": "",
                "Doc.Type": "",
                "Item.Texts": "Opening Balance",
                "Dr.Amount":  0,
                "Cr.Amount": 0,
                "Balance": opening_balance
            })

            for entry in all_entries:
                entry['Balance'] = running_balance
                if entry["Dr.Amount"]:
                    running_balance += entry["Dr.Amount"]
                if entry["Cr.Amount"]:
                    running_balance -= entry["Cr.Amount"]
                entry["Balance"] = running_balance
                ledger_entries.append(entry)

            workbook = xlwt.Workbook(encoding='utf-8')
            worksheet = workbook.add_sheet('Ledger')
            title_style = xlwt.easyxf('font: bold on, height 240; align: horiz left')
            final_balance_style = xlwt.easyxf('font: bold on; align: horiz right', num_format_str='#,##0.00')
            worksheet.write_merge(0, 0, 0, 6, f"Company: {company_name}", title_style)
            worksheet.write_merge(1, 1, 0, 6, f"Account Statement from {start_date.strftime('%Y-%m-%d')} to {end_date.strftime('%Y-%m-%d')}")
            headers = ["Doc.No", "Doc.Date", "Doc.Type", "Item.Texts", "Dr.Amount", "Cr.Amount", "Balance"]
            header_style = xlwt.easyxf('font: bold on; align: horiz center')
            number_style = xlwt.easyxf(num_format_str='#,##0.00')
            date_style = xlwt.XFStyle()
            date_style.num_format_str = 'yyyy-mm-dd'


            for col_num, header in enumerate(headers):
                worksheet.write(2, col_num, header, header_style)


            row_num = 3
            for entry in ledger_entries:
                worksheet.write(row_num, 0, entry["Doc.No"])
                if isinstance(entry["Doc.Date"], date):
                    worksheet.write(row_num, 1, entry["Doc.Date"], date_style) 
                else:
                     worksheet.write(row_num, 1, entry["Doc.Date"] )
                worksheet.write(row_num, 2, entry["Doc.Type"])
                worksheet.write(row_num, 3, entry["Item.Texts"])
                worksheet.write(row_num, 4, entry["Dr.Amount"], number_style)
                worksheet.write(row_num, 5, entry["Cr.Amount"], number_style)
                worksheet.write(row_num, 6, entry["Balance"], number_style)

                row_num += 1
            worksheet.write(row_num, 0, "Final Balance:", xlwt.easyxf('font: bold on; align: horiz right'))
            worksheet.write(row_num, 6, ledger_entries[-1]['Balance'], final_balance_style)
            worksheet.col(0).width=80*50
            worksheet.col(1).width=80*50
            worksheet.col(3).width = 256 * 50 
            today=date.today()
            response = HttpResponse(content_type='application/ms-excel')
            response['Content-Disposition'] = f'attachment; filename="{company_name}_{today}_ledger.xls"'
            workbook.save(response)
            return response

        except Exception as e:
            return JsonResponse({'error_message': str(e)}, status=500)

    return JsonResponse({'error': 'Invalid request method. Use GET.'}, status=405)

def safe_float(value):
    try:
        if value is None or value == '':
            return 0.0
        return float(value)
    except (ValueError, TypeError):
        return 0.0


@csrf_exempt
def import_excel_data(request):
    if request.method == 'POST' and request.FILES.get('excelFile'):
        excel_file = request.FILES['excelFile']
        try:
            wb = load_workbook(excel_file)
            sheet = wb.active
            errors = []
            payment_method = "default"

            with transaction.atomic():
                for index, row in enumerate(sheet.iter_rows(min_row=2, values_only=True), start=2):
                    try:
                        customer_name = str(row[2]).strip() if row[2] else None
                        customer_area = str(row[5]).strip() if row[5] else ''  
                        customer_phone = str(row[3]).strip() if row[3] else ''  
                        product = str(row[6]).strip() if row[6] else ''  
                        billing_date = row[8] 
                        updated_time = row[9] 
                        amount = row[10] 
                        total_paid = row[11]  
                        balance = row[12] if len(row) > 12 else None  
                        status = str(row[13]).strip().lower() if len(row) > 13 and row[13] else 'unpaid' 
                        remarks = str(row[14]).strip() if len(row) > 14 and row[14] else ""  


                        if not customer_name:
                            raise ValueError("Customer name is required")

                        if billing_date:
                            if isinstance(billing_date, float) or isinstance(billing_date, int):

                                try:
                                    billing_date = opxldatetime.from_excel(billing_date).date()
                                except Exception as e:
                                    errors.append(f"Row {index}: Could not convert excel number to date: {str(e)}")
                                    billing_date = None
                            elif isinstance(billing_date, str):
                                billing_date = billing_date.strip()
                                if billing_date:
                                    for fmt in ('%Y-%m-%d %H:%M:%S', '%d/%m/%Y', '%Y-%m-%d'):
                                        try:
                                            billing_date = datetime.strptime(billing_date, fmt).date()
                                            break
                                        except ValueError:
                                            continue
                                    else:
                                        errors.append(f"Row {index}: Invalid date format for 'billing_date': {billing_date}. Expecting %Y-%m-%d %H:%M:%S, %d/%m/%Y or %Y-%m-%d")
                                        billing_date = None
                                else:
                                    billing_date = None  
                            elif isinstance(billing_date, datetime):
                                billing_date = billing_date.date()
                            else:
                                errors.append(f"Row {index}: Invalid data type for 'billing_date': {billing_date}. Should be a string or datetime object.")
                                billing_date = None
                        else:
                            billing_date = None

                        if updated_time:
                            if isinstance(updated_time, float) or isinstance(updated_time, int):

                                try:
                                    updated_time = opxldatetime.from_excel(updated_time).date() 
                                except Exception as e:
                                    errors.append(f"Row {index}: Could not convert excel number to date: {str(e)}")
                                    updated_time = None
                            elif isinstance(updated_time, str):
                                updated_time = updated_time.strip()
                                if updated_time:
                                     date_part=updated_time.split(" ")[0]
                                     for fmt in ('%Y-%m-%d', '%d/%m/%Y', '%#d/%#m/%Y'):
                                        try:
                                             updated_time=datetime.strptime(date_part, fmt).date()
                                             break
                                        except ValueError:
                                            continue
                                     else:
                                        errors.append(f"Row {index}: Invalid datetime format for 'updated_time': {updated_time}")
                                        updated_time = None
                                else:
                                    updated_time= None
                            elif isinstance(updated_time, datetime):
                                updated_time = updated_time.date() 
                            else:
                                errors.append(f"Row {index}: Invalid data type for 'updated_time': {updated_time}. Should be a string or datetime object.")
                                updated_time = None
                        else:
                            updated_time = None

                        amount = safe_float(amount)
                        total_paid = safe_float(total_paid)
                        if billing_date is None:
                            billing_date = timezone.now().date() 

                        customer, created = Customer.objects.get_or_create(name=customer_name, area=customer_area)
                        order, created = Order.objects.update_or_create(
                            bill_no=row[1],
                            defaults={
                                'customer': customer,
                                'product': product,
                                'billing_date': billing_date,
                                'amount': amount
                            }
                        )
                        if updated_time and remarks and total_paid != Decimal(0.00):
                            payment_amount =total_paid
                            if payment_amount != Decimal(0.00):
                                Payment.objects.create(
                                    order=order,
                                    payment_date=updated_time,
                                    paid_amount=payment_amount, 
                                    remarks=remarks,
                                    payment_method=payment_method,
                                )

                    except Exception as e:
                        errors.append(f"Row {index}: {str(e)}")

            if errors:
                return JsonResponse({'success': False, 'error_message': 'Errors during import', 'errors': errors}, status=400)

            return JsonResponse({'success': True, 'message': 'Data imported successfully'}, status=201)

        except Exception as e:
            return JsonResponse({'success': False, 'error_message': f"Error processing file: {str(e)}"}, status=400)

    return JsonResponse({'success': False, 'error_message': 'Invalid request'}, status=400)